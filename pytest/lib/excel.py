
#!/usr/bin/python3
# -*- coding:utf-8 -*-

import os
import sys
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
print(BASE_DIR)
sys.path.append(BASE_DIR)
from config import excelConfig


class excel:
    def __init__(self, excelConfig):
        self.workbook = os.path.join('testcase', 'resource', excelConfig['workbook'])
        self.rtsheet = excelConfig['rtsheet']
        self.sheets = excelConfig['sheets']
        self.driverfuncsheet=excelConfig['driverfuncsheet']
        self.basicfuncsheet=excelConfig['basicfuncsheet']
        self.perfsheet=excelConfig['perfsheet']
        self.relisheet=excelConfig['relisheet']
 
    def get_test_case(self):
        wb = load_workbook(filename=self.workbook)  # 实例化excel
        #sh = wb[self.sheet]  # 获取sheet
        test_case_list = []  # 定义一个全局变量存放
        for sheet_name in self.sheets:
         sh = wb[sheet_name]  # 获取sheet
         all_excel_data = list(sh.iter_rows(values_only=True))  # 获取表格所有数据
         excel_title = [t.strip() if t else t for t in all_excel_data[0]]  # 获取表头
            # excel_title_index = [excel_title.index(c) for c in all_excel_data[0]]  # 获取表头字母
         case_data_list = all_excel_data[1:]  # 获取用例数据
         for case in case_data_list:
                # print('表头', list(excel_title))
                # print('测试用例', case)
            res = dict(list(zip(excel_title, case)))  # 数据拼接，用表头与测试用例数据进行组装拼接成dict
            if '正则' in res and res['正则']:
                    test_case_list.append(res)
                    # print('拼接后的数据:', res)
            # print(len(test_case_list))
        wb.close()
        return test_case_list
    def get_driverfunc_testcase(self):
        wb = load_workbook(filename=self.workbook)  # 实例化excel
        test_case_list = []  # 定义一个全局变量存放
        sh = wb[self.driverfuncsheet]  # 获取sheet
        all_excel_data = list(sh.iter_rows(values_only=True))  # 获取表格所有数据
        excel_title = [t.strip() if t else t for t in all_excel_data[0]]  # 获取表头
            # excel_title_index = [excel_title.index(c) for c in all_excel_data[0]]  # 获取表头字母
        case_data_list = all_excel_data[1:]  # 获取用例数据
        for case in case_data_list:
                # print('表头', list(excel_title))
                # print('测试用例', case)
            res = dict(list(zip(excel_title, case)))  # 数据拼接，用表头与测试用例数据进行组装拼接成dict
            if '正则' in res and res['正则']:
                    test_case_list.append(res)
                    # print('拼接后的数据:', res)
                    # print(len(test_case_list))
        wb.close()
        return test_case_list
    def get_basicfunc_testcase(self):
        wb = load_workbook(filename=self.workbook)  # 实例化excel
        test_case_list = []  # 定义一个全局变量存放
        sh = wb[self.basicfuncsheet]  # 获取sheet
        all_excel_data = list(sh.iter_rows(values_only=True))  # 获取表格所有数据
        excel_title = [t.strip() if t else t for t in all_excel_data[0]]  # 获取表头
            # excel_title_index = [excel_title.index(c) for c in all_excel_data[0]]  # 获取表头字母
        case_data_list = all_excel_data[1:]  # 获取用例数据
        for case in case_data_list:
                # print('表头', list(excel_title))
                # print('测试用例', case)
            res = dict(list(zip(excel_title, case)))  # 数据拼接，用表头与测试用例数据进行组装拼接成dict
            if '正则' in res and res['正则']:
                    test_case_list.append(res)
                    # print('拼接后的数据:', res)
                    # print(len(test_case_list))
        wb.close()
        return test_case_list
    def get_rttest_case(self):
        wb = load_workbook(filename=self.workbook)  # 实例化excel
        test_case_list = []  # 定义一个全局变量存放
        sh = wb[self.rtsheet]  # 获取sheet
        all_excel_data = list(sh.iter_rows(values_only=True))  # 获取表格所有数据
        excel_title = [t.strip() if t else t for t in all_excel_data[0]]  # 获取表头
            # excel_title_index = [excel_title.index(c) for c in all_excel_data[0]]  # 获取表头字母
        case_data_list = all_excel_data[1:]  # 获取用例数据
        for case in case_data_list:
                # print('表头', list(excel_title))
                # print('测试用例', case)
            res = dict(list(zip(excel_title, case)))  # 数据拼接，用表头与测试用例数据进行组装拼接成dict
            if '正则' in res and res['正则']:
                    test_case_list.append(res)
                    # print('拼接后的数据:', res)
            # print(len(test_case_list))
        wb.close()
        return test_case_list
    def get_perftest_case(self):
        wb = load_workbook(filename=self.workbook)  # 实例化excel
        test_case_list = []  # 定义一个全局变量存放
        sh = wb[self.perfsheet]  # 获取sheet
        all_excel_data = list(sh.iter_rows(values_only=True))  # 获取表格所有数据
        excel_title = [t.strip() if t else t for t in all_excel_data[0]]  # 获取表头
            # excel_title_index = [excel_title.index(c) for c in all_excel_data[0]]  # 获取表头字母
        case_data_list = all_excel_data[1:]  # 获取用例数据
        for case in case_data_list:
                # print('表头', list(excel_title))
                # print('测试用例', case)
            res = dict(list(zip(excel_title, case)))  # 数据拼接，用表头与测试用例数据进行组装拼接成dict
            if '正则' in res and res['正则']:
                    test_case_list.append(res)
                    # print('拼接后的数据:', res)
            # print(len(test_case_list))
        wb.close()
        return test_case_list
    def get_reliability_case(self):
        wb = load_workbook(filename=self.workbook)  # 实例化excel
        test_case_list = []  # 定义一个全局变量存放
        sh = wb[self.relisheet]  # 获取sheet
        all_excel_data = list(sh.iter_rows(values_only=True))  # 获取表格所有数据
        excel_title = [t.strip() if t else t for t in all_excel_data[0]]  # 获取表头
            # excel_title_index = [excel_title.index(c) for c in all_excel_data[0]]  # 获取表头字母
        case_data_list = all_excel_data[1:]  # 获取用例数据
        for case in case_data_list:
                # print('表头', list(excel_title))
                # print('测试用例', case)
            res = dict(list(zip(excel_title, case)))  # 数据拼接，用表头与测试用例数据进行组装拼接成dict
            if '正则' in res and res['正则']:
                    test_case_list.append(res)
                    # print('拼接后的数据:', res)
                    # print(len(test_case_list))
        wb.close()
        return test_case_list
if __name__ == "__main__":
    print ('This is main of module "excel.py"')
    exl = excel(excelConfig)
    caseData = exl.get_basicfunc_testcase()
    print(caseData)